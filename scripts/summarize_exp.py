"""Summarize one experiment directory into build_stats.json + qa_analysis.xlsx.

Usage:
    python scripts/summarize_exp.py --exp-dir experiments/2026-04-27-001-refine-jump

Outputs:
    experiments/.../build/build_stats.json   — graph + trajectory stats, by sample + total
    experiments/.../qa/qa_analysis.xlsx      — QA results, 4 sheets
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
sys.path.insert(0, str(Path(__file__).parent))
from summarize_graph_structure import summarize as summarize_graph


# ---------------------------------------------------------------------------
# Build stats
# ---------------------------------------------------------------------------

def _trajectory_stats(traj_path: Path) -> dict[str, Any]:
    if not traj_path.exists():
        return {}
    trigger = skip = 0
    ops: Counter[str] = Counter()
    rejected: Counter[str] = Counter()
    failed = 0
    dedup_merges = 0

    with traj_path.open(encoding="utf-8") as f:
        for line in f:
            rec = json.loads(line)
            phase = rec.get("phase", "")
            if phase == "trigger":
                if rec.get("triggered"):
                    trigger += 1
                else:
                    skip += 1
            elif phase == "construction":
                for op in rec.get("op_log", []):
                    op_name = op.get("op", "unknown")
                    status = op.get("status", "")
                    if status == "rejected":
                        rejected[op.get("error", "rejected")] += 1
                    elif status == "error":
                        failed += 1
                    else:
                        ops[op_name] += 1
            elif phase == "dedup":
                dedup_merges += rec.get("merged", 0)

    total_batches = trigger + skip
    return {
        "total_batches": total_batches,
        "trigger_count": trigger,
        "skip_count": skip,
        "trigger_rate": round(trigger / total_batches, 3) if total_batches else None,
        "ops": dict(ops),
        "rejected_ops": dict(rejected),
        "failed_ops": failed,
        "dedup_merges": dedup_merges,
    }


def build_build_stats(exp_dir: Path) -> dict[str, Any]:
    graphs_dir = exp_dir / "build" / "graphs"
    graph_paths = sorted(graphs_dir.glob("*.json")) if graphs_dir.exists() else []

    by_sample: dict[str, Any] = {}
    for gp in graph_paths:
        sample_id = gp.stem.replace("_graph", "")
        g_stats = summarize_graph(gp)
        traj_path = exp_dir / "build" / f"graph_trajectories_{sample_id}.jsonl"
        t_stats = _trajectory_stats(traj_path)
        by_sample[sample_id] = {**g_stats, **t_stats}

    # Aggregate totals
    int_keys = [
        "node_count", "edge_count", "entity_count", "event_count",
        "invalid_edge_count", "isolated_event_count",
        "events_with_fact", "events_missing_fact",
        "events_with_quote", "events_missing_quote",
        "event_without_entity_event_edge",
        "total_batches", "trigger_count", "skip_count", "failed_ops", "dedup_merges",
    ]
    total: dict[str, Any] = {"samples": len(by_sample)}
    for key in int_keys:
        total[key] = sum(s.get(key, 0) for s in by_sample.values())

    # Aggregate ops and rejected_ops
    total_ops: Counter[str] = Counter()
    total_rejected: Counter[str] = Counter()
    for s in by_sample.values():
        total_ops.update(s.get("ops") or {})
        total_rejected.update(s.get("rejected_ops") or {})
    total["ops"] = dict(total_ops)
    total["rejected_ops"] = dict(total_rejected)

    tb = total.get("total_batches", 0)
    total["trigger_rate"] = round(total["trigger_count"] / tb, 3) if tb else None

    return {"total": total, "by_sample": by_sample}


# ---------------------------------------------------------------------------
# QA stats helpers
# ---------------------------------------------------------------------------

def _trace_stats(traces: list[dict]) -> dict[str, Any]:
    actions = [t.get("action", "") for t in traces]
    jump_count = actions.count("jump")
    raw_fb = any(a in {"raw_fallback", "answerable_refusal_raw_fallback"} for a in actions)
    forced = "forced_finish" in actions
    stop_reason = ""
    for t in traces:
        if t.get("action") == "forced_finish":
            stop_reason = t.get("reason", "")
            break
    sequence = "→".join(
        a for a in actions
        if a not in {"frontier_exhausted", "answerable_refusal_raw_fallback"}
    )
    return {
        "jump_count": jump_count,
        "raw_fallback_used": raw_fb,
        "forced_finish": forced,
        "stop_reason": stop_reason,
        "action_sequence": sequence,
    }


def load_qa_records(eval_path: Path) -> list[dict[str, Any]]:
    records = []
    if not eval_path.exists():
        return records
    with eval_path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


# ---------------------------------------------------------------------------
# XLSX builder
# ---------------------------------------------------------------------------

CAT_NAMES = {
    "1": "Cat1 单跳事实",
    "2": "Cat2 时间",
    "3": "Cat3 推理",
    "4": "Cat4 多跳",
    "5": "Cat5 对抗不可答",
}
ALL_CATS = ["1", "2", "3", "4", "5"]

_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_BLUE   = PatternFill("solid", fgColor="D9E1F2")
_GREY   = PatternFill("solid", fgColor="F2F2F2")

BOLD = Font(bold=True)


def _f1_fill(f1) -> PatternFill | None:
    if f1 is None:
        return None
    if f1 >= 0.9:
        return _GREEN
    if f1 >= 0.5:
        return _YELLOW
    return _RED


def _write_header(ws, headers: list[str], fill=None) -> None:
    fill = fill or _BLUE
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = BOLD
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center")


def _set_widths(ws, widths: list[int]) -> None:
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _pct(n, d) -> str:
    return f"{n/d*100:.1f}%" if d else "—"


def _avg(vals) -> float | None:
    v = [x for x in vals if x is not None]
    return round(sum(v) / len(v), 4) if v else None


# ── Sheet 1: Summary ────────────────────────────────────────────────────────

def _build_summary_data(
    records: list[dict],
    sample_ids: list[str],
) -> tuple[list[list], list[str]]:
    """Returns (rows, col_headers). Rows: one per category + Total row."""
    col_headers = ["Category", "题数", "Acc", "Avg F1", "Avg BLEU1", "Judge Correct"]
    for sid in sample_ids:
        col_headers += [f"{sid}\nAcc", f"{sid}\nF1"]

    def _cat_metrics(recs):
        total_q = len(recs)
        judge_correct = sum(1 for r in recs if r.get("judge_label") == "CORRECT")
        f1s = [r.get("f1_score") for r in recs if r.get("f1_score") is not None]
        b1s = [r.get("bleu1_score") for r in recs if r.get("bleu1_score") is not None]
        acc = _pct(judge_correct, total_q)
        return {
            "n": total_q,
            "acc": acc,
            "f1": _avg(f1s),
            "bleu1": _avg(b1s),
            "judge": f"{judge_correct}/{total_q}",
        }

    rows = []
    all_cats_in_data = sorted({str(r.get("category", "")) for r in records})
    cats_to_show = [c for c in ALL_CATS if c in all_cats_in_data]

    for cat in cats_to_show:
        cat_recs = [r for r in records if str(r.get("category")) == cat]
        m = _cat_metrics(cat_recs)
        row = [CAT_NAMES.get(cat, f"Cat{cat}"), m["n"], m["acc"], m["f1"], m["bleu1"], m["judge"]]
        for sid in sample_ids:
            s_recs = [r for r in cat_recs if r.get("sample_id") == sid]
            sm = _cat_metrics(s_recs)
            row += [sm["acc"], sm["f1"]]
        rows.append(row)

    # Total row (Cat1-4 + All)
    for label, cat_filter in [("Cat1-4", lambda r: str(r.get("category")) in {"1","2","3","4"}),
                               ("Total (All)", lambda r: True)]:
        sub = [r for r in records if cat_filter(r)]
        m = _cat_metrics(sub)
        row = [label, m["n"], m["acc"], m["f1"], m["bleu1"], m["judge"]]
        for sid in sample_ids:
            s_recs = [r for r in sub if r.get("sample_id") == sid]
            sm = _cat_metrics(s_recs)
            row += [sm["acc"], sm["f1"]]
        rows.append(row)

    return rows, col_headers


def _write_summary_sheet(wb: openpyxl.Workbook, records: list[dict], sample_ids: list[str]) -> None:
    ws = wb.create_sheet("Summary")
    rows, headers = _build_summary_data(records, sample_ids)
    _write_header(ws, headers)
    for i, row in enumerate(rows, 2):
        fill = _GREY if row[0].startswith("Cat1-4") or row[0].startswith("Total") else None
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if fill:
                c.fill = fill
            if col == 1 and fill:
                c.font = BOLD
    widths = [18, 6, 8, 8, 8, 12] + [8, 8] * len(sample_ids)
    _set_widths(ws, widths)
    ws.freeze_panes = "A2"


# ── Sheet 2: By Conversation ─────────────────────────────────────────────────

def _write_by_conv_sheet(wb: openpyxl.Workbook, records: list[dict], sample_ids: list[str]) -> None:
    ws = wb.create_sheet("By Conversation")
    cats = [c for c in ALL_CATS if any(str(r.get("category")) == c for r in records)]
    headers = ["sample_id", "总题数", "总Acc", "总F1"]
    for cat in cats:
        headers += [f"Cat{cat}\n题数", f"Cat{cat}\nAcc", f"Cat{cat}\nF1"]
    headers += ["raw_fallback率", "forced_finish率", "avg_hops"]
    _write_header(ws, headers)

    for row_i, sid in enumerate(sample_ids, 2):
        s_recs = [r for r in records if r.get("sample_id") == sid]
        total_q = len(s_recs)
        judge_correct = sum(1 for r in s_recs if r.get("judge_label") == "CORRECT")
        f1s = [r.get("f1_score") for r in s_recs if r.get("f1_score") is not None]

        ts_list = [_trace_stats(r.get("traces") or []) for r in s_recs]
        rfb_rate = _pct(sum(1 for t in ts_list if t["raw_fallback_used"]), total_q)
        ff_rate  = _pct(sum(1 for t in ts_list if t["forced_finish"]), total_q)
        avg_hops = _avg([t["jump_count"] for t in ts_list])

        row = [sid, total_q, _pct(judge_correct, total_q), _avg(f1s)]
        for cat in cats:
            cr = [r for r in s_recs if str(r.get("category")) == cat]
            cj = sum(1 for r in cr if r.get("judge_label") == "CORRECT")
            cf1 = [r.get("f1_score") for r in cr if r.get("f1_score") is not None]
            row += [len(cr), _pct(cj, len(cr)), _avg(cf1)]
        row += [rfb_rate, ff_rate, avg_hops]

        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="center")

    widths = [14, 6, 8, 8] + [6, 8, 8] * len(cats) + [10, 10, 8]
    _set_widths(ws, widths)
    ws.freeze_panes = "A2"


# ── Sheet 3: All Questions ────────────────────────────────────────────────────

Q_HEADERS = [
    "#", "qa_id", "sample_id", "category",
    "question", "gold", "pred",
    "judge", "judge_reasoning",
    "f1", "bleu1",
    "hops", "action_sequence", "raw_fallback", "forced_finish", "stop_reason",
]
Q_WIDTHS = [4, 18, 10, 6, 45, 30, 30, 8, 35, 6, 6, 5, 30, 10, 10, 14]


def _write_questions_sheet(
    wb: openpyxl.Workbook,
    records: list[dict],
    sheet_name: str = "All Questions",
    filter_wrong: bool = False,
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, Q_HEADERS)

    row_i = 2
    for rec in records:
        judge = rec.get("judge_label", "")
        if filter_wrong and judge != "WRONG":
            continue
        f1 = rec.get("f1_score")
        ts = _trace_stats(rec.get("traces") or [])
        row = [
            row_i - 1,
            rec.get("qa_id", ""),
            rec.get("sample_id", ""),
            rec.get("category", ""),
            rec.get("question", ""),
            str(rec.get("gold", "")),
            rec.get("pred", ""),
            judge,
            rec.get("judge_reasoning", ""),
            round(f1, 3) if isinstance(f1, float) else f1,
            round(rec.get("bleu1_score", 0) or 0, 3),
            ts["jump_count"],
            ts["action_sequence"],
            "Y" if ts["raw_fallback_used"] else "",
            "Y" if ts["forced_finish"] else "",
            ts["stop_reason"],
        ]
        fill = _f1_fill(f1)
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill and col in (6, 7, 10):
                c.fill = fill
            if judge == "WRONG" and col in (5, 6, 7):
                c.fill = _RED
        row_i += 1

    _set_widths(ws, Q_WIDTHS)
    ws.freeze_panes = "E2"


# ── Sheet 4: By Category ─────────────────────────────────────────────────────

def _write_by_category_sheets(wb: openpyxl.Workbook, records: list[dict]) -> None:
    by_cat: dict[str, list] = defaultdict(list)
    for r in records:
        by_cat[str(r.get("category", "?"))].append(r)

    for cat in sorted(by_cat.keys()):
        cat_recs = by_cat[cat]
        ws = wb.create_sheet(CAT_NAMES.get(cat, f"Cat{cat}"))
        _write_header(ws, Q_HEADERS)
        row_i = 2
        for rec in cat_recs:
            f1 = rec.get("f1_score")
            judge = rec.get("judge_label", "")
            ts = _trace_stats(rec.get("traces") or [])
            row = [
                row_i - 1,
                rec.get("qa_id", ""),
                rec.get("sample_id", ""),
                cat,
                rec.get("question", ""),
                str(rec.get("gold", "")),
                rec.get("pred", ""),
                judge,
                rec.get("judge_reasoning", ""),
                round(f1, 3) if isinstance(f1, float) else f1,
                round(rec.get("bleu1_score", 0) or 0, 3),
                ts["jump_count"],
                ts["action_sequence"],
                "Y" if ts["raw_fallback_used"] else "",
                "Y" if ts["forced_finish"] else "",
                ts["stop_reason"],
            ]
            fill = _f1_fill(f1)
            for col, val in enumerate(row, 1):
                c = ws.cell(row=row_i, column=col, value=val)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if fill and col in (6, 7, 10):
                    c.fill = fill
                if judge == "WRONG" and col in (5, 6, 7):
                    c.fill = _RED
            row_i += 1
        _set_widths(ws, Q_WIDTHS)
        ws.freeze_panes = "E2"


def build_qa_xlsx(eval_path: Path, output_path: Path) -> None:
    records = load_qa_records(eval_path)
    if not records:
        print(f"  No records found in {eval_path}")
        return

    sample_ids = sorted({r.get("sample_id", "") for r in records})
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, records, sample_ids)
    _write_by_conv_sheet(wb, records, sample_ids)
    _write_questions_sheet(wb, records, sheet_name="All Questions")
    _write_questions_sheet(wb, records, sheet_name="Failures", filter_wrong=True)
    _write_by_category_sheets(wb, records)

    wb.save(output_path)
    print(f"  qa_analysis.xlsx saved: {output_path}  ({len(records)} records, {len(sample_ids)} samples)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--exp-dir", required=True, help="Experiment directory")
    p.add_argument("--skip-build", action="store_true", help="Skip build_stats.json")
    p.add_argument("--skip-qa",    action="store_true", help="Skip qa_analysis.xlsx")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    exp_dir = Path(args.exp_dir)

    if not args.skip_build:
        print("Building build_stats.json …")
        stats = build_build_stats(exp_dir)
        out = exp_dir / "build" / "build_stats.json"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        total = stats["total"]
        print(f"  nodes={total.get('node_count')}  edges={total.get('edge_count')}"
              f"  trigger_rate={total.get('trigger_rate')}"
              f"  samples={total.get('samples')}")
        print(f"  Saved: {out}")

    if not args.skip_qa:
        print("Building qa_analysis.xlsx …")
        eval_path = exp_dir / "qa" / "qa_results_eval.jsonl"
        if not eval_path.exists():
            print(f"  Not found: {eval_path}  — skipping xlsx")
        else:
            out = exp_dir / "qa" / "qa_analysis.xlsx"
            build_qa_xlsx(eval_path, out)


if __name__ == "__main__":
    main()
