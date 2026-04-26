"""Export QA results to xlsx with one sheet per category.

Usage:
    python scripts/export_qa_xlsx.py \
        --input runs/qa_dashscope_v2/qa_results_eval.jsonl \
        --output runs/qa_dashscope_v2/qa_review.xlsx
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CAT_NAMES = {
    "1": "Cat1 单跳事实",
    "2": "Cat2 时间",
    "3": "Cat3 开放推理",
    "4": "Cat4 多跳",
    "5": "Cat5 对抗不可答",
}
HEADERS = ["#", "sample_id", "question", "gold", "pred", "f1", "judge"]


def _cell_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _f1_fill(f1) -> PatternFill | None:
    if f1 is None:
        return None
    if f1 >= 0.9:
        return _cell_fill("C6EFCE")   # green
    if f1 >= 0.5:
        return _cell_fill("FFEB9C")   # yellow
    return _cell_fill("FFC7CE")       # red


def build_xlsx(input_path: Path, output_path: Path) -> None:
    records: dict[str, list] = defaultdict(list)
    with input_path.open(encoding="utf-8") as f:
        for line in f:
            r = json.loads(line)
            cat = str(r.get("category", "?"))
            records[cat].append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for cat in sorted(records.keys()):
        sheet_name = CAT_NAMES.get(cat, f"Cat{cat}")
        ws = wb.create_sheet(title=sheet_name)
        rows = records[cat]

        # Header row
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = _cell_fill("D9E1F2")
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Data rows
        for i, r in enumerate(rows, 1):
            f1 = r.get("f1_score")
            judge = r.get("judge_label", "")
            row_data = [
                i,
                r.get("sample_id", ""),
                r.get("question", ""),
                r.get("gold", ""),
                r.get("pred", ""),
                round(f1, 3) if isinstance(f1, float) else f1,
                judge,
            ]
            fill = _f1_fill(f1)
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill and col in (3, 4, 5, 6):
                    cell.fill = fill

        # Column widths
        widths = [5, 10, 45, 30, 30, 7, 10]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

        print(f"  Sheet '{sheet_name}': {len(rows)} rows")

    wb.save(output_path)
    print(f"Saved: {output_path}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--input",  default="runs/qa_dashscope_v2/qa_results_eval.jsonl")
    p.add_argument("--output", default="runs/qa_dashscope_v2/qa_review.xlsx")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_xlsx(Path(args.input), Path(args.output))
